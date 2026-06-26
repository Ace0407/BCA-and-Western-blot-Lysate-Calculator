#!/usr/bin/env python3
"""
BCA Protein Assay Calculator
Reads plate reader absorbance data, fits a linear standard curve,
calculates unknown protein concentrations, and exports results.

Usage:
    python bca_calculator.py input_file.xlsx
    python bca_calculator.py input_file.csv --output results.xlsx --cv-threshold 10
"""

import sys
import os
import re
import shutil
import argparse
import warnings
from datetime import datetime

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")  # non-interactive backend; safe for all terminals
import matplotlib.pyplot as plt
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Constants ────────────────────────────────────────────────────────────────

DEFAULT_CV_THRESHOLD  = 15.0    # % CV above which replicates are flagged as high-variability
NEAR_UPPER_LIMIT_CONC = 1500.0  # Pierce manual notes curve may deviate from linear above this
FIT_LINEAR = "linear"
FIT_PTP    = "point-to-point"   # piecewise interpolation — Pierce protocol recommendation
PLATE_ROWS = list("ABCDEFGH")


# ═══════════════════════════════════════════════════════════════════════════════
# DATA LOADING
# ═══════════════════════════════════════════════════════════════════════════════

def _try_parse_softmax(filepath: str) -> "pd.DataFrame | None":
    """
    Try to read a SoftMax Pro UTF-16 export (.txt or .xls from the plate reader).
    Returns a plate-grid DataFrame (row label in col 0, absorbances in cols 1-12)
    or None if the file is not in SoftMax Pro format.
    """
    try:
        with open(filepath, encoding="utf-16") as f:
            content = f.read()
    except (UnicodeDecodeError, UnicodeError, OSError):
        return None

    if not any(m in content for m in ("##BLOCKS", "~End", "Plate:\t")):
        return None

    data_rows = []
    for line in content.splitlines():
        if not line.startswith("\t\t"):
            continue
        parts = line.split("\t")
        try:
            int(parts[2])   # column-number header line — skip
            continue
        except (ValueError, IndexError):
            pass
        data_rows.append(parts[2:])

    if not data_rows:
        return None

    rows_out = []
    for i, row in enumerate(data_rows):
        if i >= len(PLATE_ROWS):
            break
        label = PLATE_ROWS[i]
        padded = (row + [""] * 12)[:12]
        rows_out.append([label] + padded)

    return pd.DataFrame(rows_out)


def load_plate_data(filepath: str, sheet: str | None = None) -> dict[str, float]:
    """
    Load absorbance values from a plate reader file.

    Supported inputs (all auto-detected):
      • SoftMax Pro UTF-16 export  — .txt or .xls directly from the plate reader
      • Real Excel file            — .xlsx or .xls saved in Excel
      • CSV                        — .csv with row labels

    Within each file, two plate layouts are auto-detected:
      1. Plate-grid  — rows labelled A–H, columns 1–12
      2. Tabular     — columns named 'Well' and 'Absorbance'

    Returns dict mapping well ID (e.g. 'A1') to absorbance float.
    """
    # SoftMax Pro detection runs first — handles .txt and .xls plate reader exports
    softmax_df = _try_parse_softmax(filepath)
    if softmax_df is not None:
        print("  [SoftMax Pro] UTF-16 format detected — converting automatically.")
        return _parse_plate_grid(softmax_df)

    ext = os.path.splitext(filepath)[1].lower()

    if ext in (".xlsx", ".xls"):
        raw = pd.read_excel(filepath, sheet_name=sheet or 0, header=None)
    elif ext == ".csv":
        raw = pd.read_csv(filepath, header=None)
    elif ext == ".txt":
        raise ValueError(
            "This .txt file was not recognised as a SoftMax Pro export. "
            "Check that the file is the direct plate reader output."
        )
    else:
        raise ValueError(
            f"Unsupported format '{ext}'. "
            "Use .txt or .xls (SoftMax Pro), .xlsx, or .csv."
        )

    # Remove fully empty rows/columns
    raw = raw.dropna(how="all").dropna(axis=1, how="all").reset_index(drop=True)

    # Detect plate-grid format: first column contains letters A–H
    first_col = [str(v).strip().upper() for v in raw.iloc[:, 0]]
    if any(r in first_col for r in PLATE_ROWS):
        return _parse_plate_grid(raw)
    else:
        return _parse_tabular(raw)


def _parse_plate_grid(df: pd.DataFrame) -> dict[str, float]:
    """Parse a plate reader grid where rows are A–H and columns are 1–N."""
    well_data: dict[str, float] = {}
    for _, row in df.iterrows():
        row_label = str(row.iloc[0]).strip().upper()
        if row_label not in PLATE_ROWS:
            continue
        for col_idx, val in enumerate(row.iloc[1:], start=1):
            try:
                well_data[f"{row_label}{col_idx}"] = float(val)
            except (ValueError, TypeError):
                pass
    return well_data


def _parse_tabular(df: pd.DataFrame) -> dict[str, float]:
    """
    Parse a two-column tabular file.
    Tries to find 'Well' and 'Absorbance' columns by name; falls back to
    first two columns positionally.
    """
    # Treat first row as header
    df.columns = [str(c).strip() for c in df.iloc[0]]
    df = df[1:].reset_index(drop=True)

    well_col = abs_col = None
    for col in df.columns:
        lc = col.lower()
        if "well" in lc:
            well_col = col
        if any(k in lc for k in ("abs", "od", "value", "reading")):
            abs_col = col

    if well_col is None or abs_col is None:
        cols = df.columns.tolist()
        if len(cols) >= 2:
            well_col, abs_col = cols[0], cols[1]
        else:
            raise ValueError(
                "Cannot identify Well and Absorbance columns. "
                "Ensure headers contain 'Well' and 'Absorbance'."
            )

    well_data: dict[str, float] = {}
    for _, row in df.iterrows():
        well_id = str(row[well_col]).strip().upper()
        try:
            well_data[well_id] = float(row[abs_col])
        except (ValueError, TypeError):
            pass
    return well_data


# ═══════════════════════════════════════════════════════════════════════════════
# WELL PARSING UTILITIES
# ═══════════════════════════════════════════════════════════════════════════════

def parse_well_list(text: str) -> list[str]:
    """
    Convert a human-entered well string into a list of well IDs.

    Accepted formats
    ----------------
    - Single well  : 'A1'
    - List          : 'A1, A2, A3'  or  'A1 A2 A3'
    - Row range     : 'A1:A4'   → ['A1','A2','A3','A4']
    - Column range  : 'A1:D1'   → ['A1','B1','C1','D1']
    """
    wells: list[str] = []
    tokens = re.split(r"[,\s;]+", text.strip())

    for token in tokens:
        token = token.strip().upper()
        if not token:
            continue

        if ":" in token:
            # Range notation
            try:
                start_str, end_str = [t.strip() for t in token.split(":", 1)]
                row_s, col_s = start_str[0], int(start_str[1:])
                row_e, col_e = end_str[0], int(end_str[1:])

                if row_s == row_e:
                    # Same row, column range  A1:A4
                    for c in range(col_s, col_e + 1):
                        wells.append(f"{row_s}{c}")
                elif col_s == col_e:
                    # Same column, row range  A1:D1
                    row_s_idx = PLATE_ROWS.index(row_s)
                    row_e_idx = PLATE_ROWS.index(row_e)
                    for r in PLATE_ROWS[row_s_idx : row_e_idx + 1]:
                        wells.append(f"{r}{col_s}")
                else:
                    # Rectangular block — expand all wells row by row
                    row_s_idx = PLATE_ROWS.index(row_s)
                    row_e_idx = PLATE_ROWS.index(row_e)
                    for r in PLATE_ROWS[min(row_s_idx, row_e_idx) : max(row_s_idx, row_e_idx) + 1]:
                        for c in range(min(col_s, col_e), max(col_s, col_e) + 1):
                            wells.append(f"{r}{c}")
            except (IndexError, ValueError):
                wells.append(token)
        else:
            wells.append(token)

    return wells


def _parse_rectangle(raw_input: str):
    """
    If raw_input is a single rectangular range like 'A1:C8' (spans both rows and columns),
    return (rows, cols). Otherwise return None.
    Used by prompt_standards and prompt_unknowns to enable batch-entry mode.
    """
    tokens = re.split(r"[\s,;]+", raw_input.strip())
    if len(tokens) != 1 or ":" not in tokens[0]:
        return None
    try:
        start_str, end_str = tokens[0].upper().split(":", 1)
        row_s, col_s = start_str[0], int(start_str[1:])
        row_e, col_e = end_str[0], int(end_str[1:])
    except (ValueError, IndexError):
        return None
    if row_s == row_e or col_s == col_e:
        return None  # same row or column — not a rectangle
    if row_s not in PLATE_ROWS or row_e not in PLATE_ROWS:
        return None
    row_s_idx = PLATE_ROWS.index(row_s)
    row_e_idx = PLATE_ROWS.index(row_e)
    rows = PLATE_ROWS[min(row_s_idx, row_e_idx) : max(row_s_idx, row_e_idx) + 1]
    cols = list(range(min(col_s, col_e), max(col_s, col_e) + 1))
    return rows, cols


# ═══════════════════════════════════════════════════════════════════════════════
# TERMINAL DISPLAY
# ═══════════════════════════════════════════════════════════════════════════════

def display_plate(well_data: dict[str, float]) -> None:
    """Print the loaded plate as a formatted 96-well grid to the terminal."""
    cols = list(range(1, 13))
    print("\n" + "═" * 72)
    print("  PLATE LAYOUT — Absorbance Values")
    print("═" * 72)
    print(f"     {'':4}", end="")
    for c in cols:
        print(f"{c:>7}", end="")
    print()
    print("  " + "─" * 70)

    for row_label in PLATE_ROWS:
        print(f"  {row_label}  ", end="")
        for c in cols:
            val = well_data.get(f"{row_label}{c}")
            print(f"{val:7.4f}" if val is not None else f"{'---':>7}", end="")
        print()

    print("═" * 72 + "\n")


# ═══════════════════════════════════════════════════════════════════════════════
# INTERACTIVE WELL-ASSIGNMENT PROMPTS
# ═══════════════════════════════════════════════════════════════════════════════

def _validate_wells(user_input: str, well_data: dict) -> tuple[list[str], list[str]]:
    """Return (valid_wells, invalid_wells) for a user-entered well string."""
    requested = parse_well_list(user_input)
    valid   = [w for w in requested if w in well_data]
    invalid = [w for w in requested if w not in well_data]
    return valid, invalid


def prompt_blank_wells(well_data: dict) -> list[str]:
    """Interactively collect 0 µg/mL standard well IDs (used as the blank reference)."""
    print("─" * 60)
    print("0 µg/mL STANDARD WELLS  (blank reference)")
    print("  Enter the wells containing your 0 µg/mL BSA standard")
    print("  (the last point in the Pierce BCA standard series — Standard I).")
    print("  These are used as the blank correction reference, NOT buffer-only wells.")
    print("  Example: A9 B9 C9   or   A9:C9")

    while True:
        raw = input("\n  0 µg/mL standard wells: ").strip()
        if not raw:
            print("  [!] At least one well is required.")
            continue

        valid, invalid = _validate_wells(raw, well_data)
        if invalid:
            print(f"  [!] Not found in plate data: {', '.join(invalid)}")
        if not valid:
            print("  [!] No valid wells found. Check your plate layout above.")
            continue

        readings = "  |  ".join(f"{w}: {well_data[w]:.4f}" for w in valid)
        print(f"  Readings → {readings}")
        if input("  Confirm? [y/n]: ").strip().lower() == "y":
            return valid


def prompt_standards(well_data: dict) -> list[dict]:
    """
    Interactively collect standard wells and concentrations.

    Supports two modes:
      • One at a time : enter wells for each standard, then its concentration
      • Block entry   : enter a rectangle like 'A1:C8' — prompted once per column

    Returns list of dicts: [{'wells': [...], 'concentration': float}, ...]
    """
    print("\n" + "─" * 60)
    print("BSA STANDARDS")
    print("  Enter each standard one at a time  OR  enter a block range to add")
    print("  all columns at once (e.g. A1:C8 adds 8 standards, one per column).")
    print("  Press Enter (no wells entered) to finish.\n")

    standards: list[dict] = []
    level = 1

    while True:
        print(f"  Standard #{level}:")
        raw_wells = input(f"    Wells (e.g. A1 B1 C1  or  A1:C8 for block): ").strip()

        if not raw_wells:
            if not standards:
                print("  [!] At least one standard is required.")
                continue
            if len(standards) < 2:
                print("  [!] At least 2 standards are required for linear regression. Add another.")
                continue
            break

        # Rectangular block → batch-entry mode (one concentration prompt per column)
        rect = _parse_rectangle(raw_wells)
        if rect:
            rows, cols = rect
            print(f"  Block {raw_wells.upper()}: rows {', '.join(rows)}, {len(cols)} columns.")
            print(f"  Enter concentration for each column (press Enter to skip a column):\n")
            for col in cols:
                col_wells = [f"{r}{col}" for r in rows]
                valid_col = [w for w in col_wells if w in well_data]
                if not valid_col:
                    print(f"    Col {col}: no plate data — skipping.")
                    continue
                readings = "  |  ".join(f"{w}: {well_data[w]:.4f}" for w in valid_col)
                raw_conc = input(f"    Col {col} ({', '.join(valid_col)})  [{readings}]  µg/mL: ").strip()
                if not raw_conc:
                    print(f"    Skipped col {col}.")
                    continue
                try:
                    conc = float(raw_conc)
                except ValueError:
                    print(f"    [!] Invalid number. Skipping col {col}.")
                    continue
                standards.append({"wells": valid_col, "concentration": conc})
                print(f"    ✓ Standard #{level}: {', '.join(valid_col)} → {conc} µg/mL")
                level += 1
            continue

        # Single standard entry
        valid, invalid = _validate_wells(raw_wells, well_data)
        if invalid:
            print(f"    [!] Not found: {', '.join(invalid)}")
        if not valid:
            print("    [!] No valid wells. Try again.")
            continue

        raw_conc = input(f"    Known concentration (µg/mL): ").strip()
        try:
            conc = float(raw_conc)
        except ValueError:
            print("    [!] Invalid number. Try again.")
            continue

        standards.append({"wells": valid, "concentration": conc})
        print(f"    ✓ Standard #{level}: {', '.join(valid)} → {conc} µg/mL")
        level += 1

    return standards


def prompt_unknowns(well_data: dict) -> list[dict]:
    """
    Interactively collect unknown sample wells and names.

    Supports two modes:
      • One at a time : enter wells for each sample, then its name
      • Block entry   : enter a rectangle like 'D1:E5' — one name prompt per column

    Returns list of dicts: [{'wells': [...], 'name': str}, ...]
    Dilution factor is applied globally from the lysate:buffer ratio (set in main).
    """
    print("\n" + "─" * 60)
    print("UNKNOWN SAMPLES")
    has_blank_row = input("  Is there a blank row between the standards and your samples? [y/N]: ").strip().lower() == "y"
    if has_blank_row:
        print("  Blank row noted — samples start at row E (e.g. E1:F11).")
    else:
        print("  No blank row — samples start at row D (e.g. D1:E11).")
    print()
    print("  Enter a block range to add all samples at once, or one at a time.")
    print("  Press Enter (no wells entered) to finish.\n")

    unknowns: list[dict] = []
    num = 1

    while True:
        print(f"  Unknown #{num}:")
        raw_wells = input(f"    Wells (e.g. D1 E1  or  D1:E5 for block): ").strip()

        if not raw_wells:
            if not unknowns:
                print("  [!] At least one unknown sample is required.")
                continue
            break

        # Rectangular block → batch-entry mode (one name prompt per column)
        rect = _parse_rectangle(raw_wells)
        if rect:
            rows, cols = rect
            print(f"  Block {raw_wells.upper()}: rows {', '.join(rows)}, {len(cols)} samples.")
            want_names = input("  Name the samples? [y/N]: ").strip().lower() == "y"
            print()
            for col in cols:
                col_wells = [f"{r}{col}" for r in rows]
                valid_col = [w for w in col_wells if w in well_data]
                if not valid_col:
                    print(f"    Col {col}: no plate data — skipping.")
                    continue
                if want_names:
                    name = input(f"    Sample {num} ({', '.join(valid_col)}) name: ").strip()
                    if not name:
                        name = f"Sample_{num}"
                else:
                    name = f"Sample_{num}"
                unknowns.append({"wells": valid_col, "name": name})
                print(f"    ✓ '{name}': {', '.join(valid_col)}")
                num += 1
            continue

        # Single sample entry
        valid, invalid = _validate_wells(raw_wells, well_data)
        if invalid:
            print(f"    [!] Not found: {', '.join(invalid)}")
        if not valid:
            print("    [!] No valid wells. Try again.")
            continue

        name = input(f"    Sample name [default: Sample_{num}]: ").strip()
        if not name:
            name = f"Sample_{num}"

        unknowns.append({"wells": valid, "name": name})
        print(f"    ✓ '{name}': {', '.join(valid)}")
        num += 1

    return unknowns


def prompt_mixing_ratio() -> tuple[int, int]:
    """Ask for lysate : sample buffer mixing ratio. Returns (lysate_parts, buffer_parts)."""
    print("\n" + "─" * 60)
    print("SAMPLE BUFFER RATIO")
    print("  Enter the lysate : sample buffer mixing ratio.")
    print("  e.g. '5:1' means 5 parts lysate + 1 part sample buffer.")
    print("  This sets the concentration correction factor and loading volumes.\n")
    raw = input("  Ratio (lysate:buffer) [default: 5:1]: ").strip()
    if not raw:
        lysate_parts, buffer_parts = 5, 1
    else:
        try:
            parts = [p.strip() for p in raw.split(":")]
            lysate_parts, buffer_parts = int(parts[0]), int(parts[1])
        except (ValueError, IndexError):
            print("  [!] Invalid format. Using default 5:1.")
            lysate_parts, buffer_parts = 5, 1
    cf = (lysate_parts + buffer_parts) / lysate_parts
    print(f"  Ratio {lysate_parts}:{buffer_parts}  →  correction factor ÷{cf:.4f}")
    return lysate_parts, buffer_parts


# ═══════════════════════════════════════════════════════════════════════════════
# CALCULATIONS
# ═══════════════════════════════════════════════════════════════════════════════

def _replicate_stats(abs_values: list[float]) -> tuple[float, float]:
    """
    Return (mean, %CV) for a list of replicate absorbance values.
    %CV is 0 when there is only one replicate or the mean is zero.
    """
    mean = float(np.mean(abs_values))
    if len(abs_values) > 1 and mean != 0:
        cv = float(np.std(abs_values, ddof=1) / mean * 100)
    else:
        cv = 0.0
    return mean, cv


def compute_blank_corrected(
    well_data: dict,
    blank_wells: list[str],
    standard_defs: list[dict],
    unknown_defs: list[dict],
) -> tuple[float, float, pd.DataFrame, pd.DataFrame]:
    """
    Follow Pierce BCA protocol steps 5–6 (microplate) / 6–7 (test tube):

      1. Compute average blank absorbance from all blank wells.
      2. Subtract the average blank from EACH individual replicate.
      3. Average the blank-corrected replicate values for each group.

    CV is reported on raw replicates (blank-corrected CV is unstable near zero
    for dilute standards and is not specified by the Pierce protocol).

    Returns
    -------
    blank_avg     : float
    blank_cv      : float (%)
    standards_df  : DataFrame
    unknowns_df   : DataFrame
    """
    # ── Blank ────────────────────────────────────────────────────────────────
    blank_vals = [well_data[w] for w in blank_wells]
    blank_avg, blank_cv = _replicate_stats(blank_vals)

    # ── Standards ────────────────────────────────────────────────────────────
    std_rows = []
    for std in standard_defs:
        vals = [well_data[w] for w in std["wells"]]
        raw_avg, cv = _replicate_stats(vals)
        # Pierce protocol: subtract blank from each individual replicate, then average
        corrected = float(np.mean([v - blank_avg for v in vals]))
        std_rows.append(
            {
                "concentration_ug_mL": std["concentration"],
                "wells": ", ".join(std["wells"]),
                "n_replicates": len(vals),
                "raw_absorbance_avg": raw_avg,       # full float64; formatted at display/export only
                "blank_corrected_abs": corrected,    # used in linregress — rounding here degraded fit
                "cv_pct": round(cv, 2),              # display-only stat, 2 dp is sufficient
            }
        )

    standards_df = (
        pd.DataFrame(std_rows)
        .sort_values("concentration_ug_mL")
        .reset_index(drop=True)
    )

    # ── Unknowns ─────────────────────────────────────────────────────────────
    unk_rows = []
    for unk in unknown_defs:
        vals = [well_data[w] for w in unk["wells"]]
        raw_avg, cv = _replicate_stats(vals)
        # Pierce protocol: subtract blank from each individual replicate, then average
        corrected = float(np.mean([v - blank_avg for v in vals]))
        unk_rows.append(
            {
                "sample_name": unk["name"],
                "wells": ", ".join(unk["wells"]),
                "n_replicates": len(vals),
                "dilution_factor": unk["dilution"],
                "raw_absorbance_avg": raw_avg,       # full float64; formatted at display/export only
                "blank_corrected_abs": corrected,    # used to compute concentration — keep full precision
                "cv_pct": round(cv, 2),
            }
        )

    unknowns_df = pd.DataFrame(unk_rows)
    return blank_avg, blank_cv, standards_df, unknowns_df


def fit_linear_curve(standards_df: pd.DataFrame) -> tuple[float, float, float]:
    """
    Fit Concentration = slope × Absorbance + intercept (FORECAST direction).
    Matches Excel FORECAST(abs, known_conc, known_abs) used in the mentor template.

    Returns (slope, intercept, r_squared).
    """
    x = standards_df["blank_corrected_abs"].values.astype(float)    # absorbance on X
    y = standards_df["concentration_ug_mL"].values.astype(float)    # concentration on Y

    if len(x) < 2:
        raise ValueError("Linear regression requires at least 2 standards.")

    slope, intercept, r_value, _p, _se = stats.linregress(x, y)
    return float(slope), float(intercept), float(r_value ** 2)


def calculate_concentrations(
    unknowns_df: pd.DataFrame,
    fit_mode: str,
    standards_df: pd.DataFrame,
    slope: float,
    intercept: float,
    std_min_conc: float,
    std_max_conc: float,
    cv_threshold: float,
) -> pd.DataFrame:
    """
    Compute protein concentration for each unknown following Pierce protocol step 7.

    Linear mode  : conc_in_well = (blank_corrected_abs − intercept) / slope
    Point-to-point: piecewise interpolation between adjacent standard absorbances
                    (Pierce recommendation for hand-plotted / microplate data)

    QC flags (may be combined with ';'):
        BELOW CURVE RANGE               — concentration < lowest standard
        ABOVE CURVE RANGE               — concentration > highest standard
        NEAR UPPER CURVE LIMIT (>1500)  — linear fit less reliable near top of Pierce
                                          working range (protocol note)
        HIGH CV (xx.x%)                 — replicate variability exceeds threshold
    """
    # Pre-sort standards by blank-corrected absorbance for PTP lookup
    std_sorted   = standards_df.sort_values("blank_corrected_abs")
    ptp_abs      = std_sorted["blank_corrected_abs"].values.astype(float)
    ptp_concs    = std_sorted["concentration_ug_mL"].values.astype(float)

    results = unknowns_df.copy()
    conc_in_well_col, final_conc_col, flags_col = [], [], []

    for _, row in results.iterrows():
        y  = float(row["blank_corrected_abs"])
        cv = float(row["cv_pct"])

        # ── Concentration calculation ─────────────────────────────────────
        ciw          = np.nan
        range_status = "ok"

        if fit_mode == FIT_LINEAR:
            if slope != 0:
                ciw = slope * y + intercept   # FORECAST: conc = slope × adj_abs + intercept
            else:
                range_status = "calc_error"
        else:  # point-to-point
            if y < ptp_abs[0]:
                range_status = "below"
            elif y > ptp_abs[-1]:
                range_status = "above"
            else:
                ciw = float(np.interp(y, ptp_abs, ptp_concs))

        final = ciw / row["dilution_factor"] if not np.isnan(ciw) else np.nan

        conc_in_well_col.append(ciw   if not np.isnan(ciw)   else np.nan)   # full precision; used as neat_conc in WB loading
        final_conc_col.append(  final if not np.isnan(final) else np.nan)   # full precision

        # ── QC flags ─────────────────────────────────────────────────────
        flags: list[str] = []

        if range_status == "calc_error":
            flags.append("CALCULATION ERROR")
        elif range_status == "below":
            flags.append("BELOW CURVE RANGE")
        elif range_status == "above":
            flags.append("ABOVE CURVE RANGE")
        elif not np.isnan(ciw):
            if ciw < std_min_conc:
                flags.append("BELOW CURVE RANGE")
            elif ciw > std_max_conc:
                flags.append("ABOVE CURVE RANGE")
            elif ciw > NEAR_UPPER_LIMIT_CONC:
                # Pierce manual: curve may not be perfectly linear above 1500 µg/mL
                flags.append(f"NEAR UPPER CURVE LIMIT (>{NEAR_UPPER_LIMIT_CONC:.0f} µg/mL)")
        else:
            flags.append("CALCULATION ERROR")

        if cv > cv_threshold:
            flags.append(f"HIGH CV ({cv:.1f}%)")

        flags_col.append("; ".join(flags) if flags else "OK")

    results["concentration_in_well_ug_mL"] = conc_in_well_col
    results["final_concentration_ug_mL"]   = final_conc_col
    results["qc_flags"]                    = flags_col
    return results


# ═══════════════════════════════════════════════════════════════════════════════
# WESTERN BLOT LOADING CALCULATOR
# ═══════════════════════════════════════════════════════════════════════════════

def prompt_western_params(lysate_parts: int, buffer_parts: int) -> dict:
    """Ask user for Western blot loading parameters."""
    print("\n" + "─" * 60)
    print("WESTERN BLOT LOADING CALCULATOR")
    cf = (lysate_parts + buffer_parts) / lysate_parts
    print(f"  Protocol: {lysate_parts}:{buffer_parts} (lysate:sample buffer)  |  correction ÷{cf:.4f}\n")

    while True:
        raw = input("  Target protein per lane (µg)  [e.g. 10]: ").strip()
        try:
            target_ug = float(raw)
            break
        except ValueError:
            print("  [!] Enter a number.")

    raw = input(
        "  Max well volume (µL) — flag samples that exceed this\n"
        "  [e.g. 30, press Enter to skip]: "
    ).strip()
    max_well_vol = None
    if raw:
        try:
            max_well_vol = float(raw)
        except ValueError:
            print("  [!] Invalid. No max-volume check will be applied.")

    print(f"\n  Parameters confirmed:")
    print(f"    Target protein per lane  : {target_ug} µg")
    if max_well_vol is not None:
        print(f"    Max well volume          : {max_well_vol} µL")

    return {
        "target_ug":    target_ug,
        "max_well_vol": max_well_vol,
        "lysate_parts": lysate_parts,
        "buffer_parts": buffer_parts,
    }


def calculate_western_loading(
    unknowns_results: pd.DataFrame,
    params: dict,
) -> pd.DataFrame:
    """
    For each sample compute loading volumes using the lysate:sample buffer ratio.

        lysate_vol        = target_ug / (neat_conc_ug_mL / 1000)
        sample_buffer_vol = lysate_vol / lysate_parts
        total_load_vol    = lysate_vol + sample_buffer_vol

    neat_conc = concentration_in_well_ug_mL (FORECAST result before correction).

    WB flags:
        NO CONCENTRATION       — concentration is NaN
        INVALID CONCENTRATION  — concentration ≤ 0
        OUTSIDE CURVE RANGE    — BCA flagged sample; value may be unreliable
        EXCEEDS MAX WELL VOL   — total_load_vol > max_well_vol
    """
    target_ug    = params["target_ug"]
    lysate_parts = params["lysate_parts"]
    buffer_parts = params["buffer_parts"]
    max_well_vol = params.get("max_well_vol")

    rows = []
    for _, row in unknowns_results.iterrows():
        neat_conc = row["concentration_in_well_ug_mL"]
        buf_conc  = row["final_concentration_ug_mL"]
        bca_flags = str(row["qc_flags"])

        flags: list[str] = []
        lysate_vol = sample_buffer_vol = total_vol = np.nan

        if pd.isna(neat_conc):
            flags.append("NO CONCENTRATION")
        elif neat_conc <= 0:
            flags.append("INVALID CONCENTRATION (≤ 0)")
        else:
            lysate_vol        = target_ug / (neat_conc / 1000.0)
            sample_buffer_vol = lysate_vol / lysate_parts
            total_vol         = lysate_vol + sample_buffer_vol

            if "CURVE RANGE" in bca_flags:
                flags.append("OUTSIDE CURVE RANGE")
            if max_well_vol is not None and total_vol > max_well_vol:
                flags.append(f"EXCEEDS MAX WELL VOL ({max_well_vol:.0f} µL)")

        rows.append({
            "sample_name":                 row["sample_name"],
            "neat_conc_ug_mL":             float(neat_conc) if pd.notna(neat_conc) else np.nan,           # full precision
            "conc_in_sample_buffer_ug_mL": float(buf_conc)  if pd.notna(buf_conc)  else np.nan,           # full precision
            "target_ug":                   target_ug,
            "lysate_vol_uL":               lysate_vol        if not np.isnan(lysate_vol)        else np.nan,  # full precision
            "sample_buffer_vol_uL":        sample_buffer_vol if not np.isnan(sample_buffer_vol) else np.nan,  # full precision
            "total_load_vol_uL":           total_vol         if not np.isnan(total_vol)         else np.nan,  # full precision
            "wb_flags":                    "; ".join(flags) if flags else "OK",
        })

    return pd.DataFrame(rows)


def print_western_table(western_df: pd.DataFrame, params: dict) -> None:
    """Print the Western blot loading table to the terminal."""
    target_ug    = params["target_ug"]
    lysate_parts = params["lysate_parts"]
    buffer_parts = params["buffer_parts"]
    max_well_vol = params.get("max_well_vol")

    sep = "═" * 110
    print("\n" + sep)
    print("  WESTERN BLOT LOADING TABLE")
    hdr_line = (
        f"  Protocol: {lysate_parts}:{buffer_parts} (lysate:sample buffer)  |  Target: {target_ug} µg/lane"
    )
    if max_well_vol is not None:
        hdr_line += f"  |  Max well vol: {max_well_vol:.0f} µL"
    print(hdr_line)
    print(sep)
    print(
        f"  {'Sample':<22}  {'Neat Conc':>12}  {'Buf Conc':>12}  {'Target':>8}  "
        f"{'Lysate':>9}  {'SampBuf':>9}  {'Total':>9}  Flags"
    )
    print(
        f"  {'':22}  {'(µg/mL)':>12}  {'(µg/mL)':>12}  {'(µg)':>8}  "
        f"{'(µL)':>9}  {'(µL)':>9}  {'(µL)':>9}"
    )
    print("  " + "─" * 106)

    for _, row in western_df.iterrows():
        neat  = f"{row['neat_conc_ug_mL']:.2f}"               if pd.notna(row["neat_conc_ug_mL"])             else "N/A"
        buf   = f"{row['conc_in_sample_buffer_ug_mL']:.2f}"   if pd.notna(row["conc_in_sample_buffer_ug_mL"]) else "N/A"
        targ  = f"{row['target_ug']:.1f}"
        lys   = f"{row['lysate_vol_uL']:.2f}"                 if pd.notna(row["lysate_vol_uL"])               else "N/A"
        sbuf  = f"{row['sample_buffer_vol_uL']:.2f}"          if pd.notna(row["sample_buffer_vol_uL"])        else "N/A"
        total = f"{row['total_load_vol_uL']:.2f}"             if pd.notna(row["total_load_vol_uL"])           else "N/A"
        flag  = row["wb_flags"]
        print(
            f"  {row['sample_name']:<22}  {neat:>12}  {buf:>12}  {targ:>8}  "
            f"{lys:>9}  {sbuf:>9}  {total:>9}  {flag}"
        )

    print(sep + "\n")


# ═══════════════════════════════════════════════════════════════════════════════
# PLOTTING
# ═══════════════════════════════════════════════════════════════════════════════

def generate_standard_curve_plot(
    standards_df: pd.DataFrame,
    fit_mode: str,
    slope: float,
    intercept: float,
    r_squared: float,
    output_path: str,
) -> None:
    """
    Save a standard curve PNG.

    Linear mode      : scatter + OLS best-fit line, annotated with equation and R²
    Point-to-point   : scatter + line segments connecting adjacent standards in
                       concentration order (Pierce protocol recommendation)
    """
    x = standards_df["concentration_ug_mL"].values.astype(float)
    y = standards_df["blank_corrected_abs"].values.astype(float)

    fig, ax = plt.subplots(figsize=(8, 6))

    ax.scatter(
        x, y,
        color="steelblue", s=80, zorder=5, edgecolors="navy", linewidths=0.8,
        label="BSA standards",
    )

    if fit_mode == FIT_LINEAR:
        # Regression: Conc = slope × Abs + intercept  →  invert for plot (Conc on X, Abs on Y)
        x_fit = np.linspace(0, x.max() * 1.1, 300)
        y_fit = (x_fit - intercept) / slope
        ax.plot(x_fit, y_fit, color="crimson", linewidth=2, zorder=4, label="Linear fit (FORECAST)")
        sign  = "+" if intercept >= 0 else "−"
        annot = (
            f"Conc = {slope:.2f} × Adj.Abs {sign} {abs(intercept):.2f}\n"
            f"$R^2$ = {r_squared:.5f}"
        )
    else:
        # Point-to-point: connect standards sorted by concentration
        idx_sort = np.argsort(x)
        ax.plot(
            x[idx_sort], y[idx_sort],
            color="crimson", linewidth=2, zorder=4,
            label="Point-to-point (Pierce protocol)",
        )
        annot = "Point-to-point interpolation\n(Pierce BCA protocol recommendation)"

    ax.annotate(
        annot,
        xy=(0.05, 0.93), xycoords="axes fraction",
        fontsize=11, verticalalignment="top",
        bbox=dict(boxstyle="round,pad=0.4", facecolor="lightyellow", edgecolor="gray", alpha=0.9),
    )

    mode_label = "Linear Regression" if fit_mode == FIT_LINEAR else "Point-to-Point"
    ax.set_xlabel("BSA Concentration (µg/mL)", fontsize=13)
    ax.set_ylabel("Blank-Corrected Absorbance (562 nm)", fontsize=13)
    ax.set_title(
        f"BCA Protein Assay — Standard Curve ({mode_label})",
        fontsize=14, fontweight="bold",
    )
    ax.legend(fontsize=11, loc="lower right")
    ax.grid(True, linestyle="--", alpha=0.4)
    ax.set_xlim(left=0)
    ax.set_ylim(bottom=0)

    plt.tight_layout()
    plt.savefig(output_path, dpi=150, bbox_inches="tight")
    plt.close()

    print(f"  [+] Standard curve plot saved → {output_path}")


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ═══════════════════════════════════════════════════════════════════════════════

# ── Style helpers ────────────────────────────────────────────────────────────

_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_HDR_FILL  = PatternFill("solid", fgColor="2E4057")
_HDR_FONT  = Font(color="FFFFFF", bold=True)
_WARN_FILL = PatternFill("solid", fgColor="FFD166")
_FLAG_FILL = PatternFill("solid", fgColor="FFBABA")
_OK_FILL   = PatternFill("solid", fgColor="C8F5C8")
_NUM_FMT   = "0.000000000"   # 9 decimal places for all numeric data cells in export


def _header_row(ws, row_num: int, headers: list[str]) -> None:
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=text)
        cell.fill   = _HDR_FILL
        cell.font   = _HDR_FONT
        cell.border = _THIN
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row_num].height = 36


def _data_cell(ws, row: int, col: int, value, flag: bool = False, num_format: str = None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.border    = _THIN
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if flag:
        cell.fill = _FLAG_FILL
    if num_format and isinstance(value, float) and not np.isnan(value):
        cell.number_format = num_format


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = max(
            (len(str(c.value)) for c in col if c.value is not None),
            default=8,
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 45)


def export_to_excel(
    output_path: str,
    blank_avg: float,
    blank_cv: float,
    blank_wells: list[str],
    standards_df: pd.DataFrame,
    unknowns_results: pd.DataFrame,
    fit_mode: str,
    slope: float,
    intercept: float,
    r_squared: float,
    cv_threshold: float,
    input_filename: str,
    western_df: pd.DataFrame | None = None,
    western_params: dict | None = None,
) -> None:
    """Write a formatted multi-sheet Excel workbook with all results."""
    wb = Workbook()

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"

    ws.merge_cells("A1:C1")
    ws["A1"].value     = "BCA Protein Assay — Results Summary"
    ws["A1"].font      = Font(size=15, bold=True, color="2E4057")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    if fit_mode == FIT_LINEAR:
        sign = "+" if (not np.isnan(intercept) and intercept >= 0) else "−"
        curve_rows = [
            ("Fit mode",    "Linear regression — FORECAST direction (Conc = f(Abs))"),
            ("Slope",       slope     if not np.isnan(slope)     else "N/A"),   # full precision; cell formatted at 9 dp
            ("Intercept",   intercept if not np.isnan(intercept) else "N/A"),   # full precision
            ("R²",          r_squared if not np.isnan(r_squared) else "N/A"),   # full precision
            ("Equation",    f"Conc = {slope:.4f} × Adj.Abs {sign} {abs(intercept):.4f}"
                            if not np.isnan(slope) else "N/A"),
        ]
    else:
        curve_rows = [
            ("Fit mode",    "Point-to-point interpolation (Pierce protocol)"),
            ("Slope",       "N/A — point-to-point has no single slope"),
            ("Intercept",   "N/A"),
            ("R²",          "N/A"),
            ("Equation",    "Piecewise linear between adjacent standards"),
        ]

    info_rows = [
        ("Input file",                    input_filename),
        ("0 µg/mL standard wells",        ", ".join(blank_wells)),
        ("0 µg/mL std avg absorbance",    blank_avg),   # full precision; cell formatted at 9 dp
        ("0 µg/mL std CV (%)",            round(blank_cv, 2)),
        ("",                      ""),
        ("─── Standard Curve ───",""),
        *curve_rows,
        ("CV flag threshold (%)", cv_threshold),
        ("",                      ""),
        ("─── Protocol Note ───", ""),
        ("Pierce BCA manual note",
         "The Pierce BCA manual states that best-fit or point-to-point curves "
         "may be more accurate than a purely linear fit for microplate data. "
         "This calculator uses linear regression by default. "
         "Run with --fit-mode point-to-point to use point-to-point interpolation."),
    ]

    for i, (label, value) in enumerate(info_rows, start=3):
        lbl_cell = ws.cell(row=i, column=1, value=label)
        lbl_cell.font = Font(bold=True)
        val_cell = ws.cell(row=i, column=2, value=value)
        if isinstance(value, float) and not np.isnan(value):
            val_cell.number_format = _NUM_FMT
        if label == "Pierce BCA manual note":
            val_cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[i].height = 60
            ws.column_dimensions["B"].width = 80

    _auto_width(ws)

    # ── Sheet 2: Standards ───────────────────────────────────────────────────
    ws_std = wb.create_sheet("Standards")
    std_headers = [
        "Concentration (µg/mL)", "Wells", "N Replicates",
        "Raw Avg Absorbance", "Blank-Corrected Abs", "CV (%)",
    ]
    _header_row(ws_std, 1, std_headers)

    for ri, row in standards_df.iterrows():
        r = ri + 2
        vals = [
            row["concentration_ug_mL"], row["wells"], row["n_replicates"],
            row["raw_absorbance_avg"], row["blank_corrected_abs"], row["cv_pct"],
        ]
        high_cv = float(row["cv_pct"]) > cv_threshold
        for ci, val in enumerate(vals, 1):
            # cols 4 (raw abs) and 5 (corrected abs) carry full float precision
            nf = _NUM_FMT if ci in (4, 5) else None
            _data_cell(ws_std, r, ci, val, flag=(ci == 6 and high_cv), num_format=nf)

    _auto_width(ws_std)

    # ── Sheet 3: Unknowns ────────────────────────────────────────────────────
    ws_unk = wb.create_sheet("Unknowns")
    unk_headers = [
        "Sample Name", "Wells", "N Replicates", "Dilution Factor",
        "Raw Avg Abs", "Blank-Corrected Abs", "CV (%)",
        "Conc. in Well (µg/mL)", "Final Concentration (µg/mL)", "QC Flags",
    ]
    _header_row(ws_unk, 1, unk_headers)

    for ri, row in unknowns_results.iterrows():
        r = ri + 2
        flagged = row["qc_flags"] != "OK"
        high_cv = float(row["cv_pct"]) > cv_threshold
        vals = [
            row["sample_name"], row["wells"], row["n_replicates"], row["dilution_factor"],
            row["raw_absorbance_avg"], row["blank_corrected_abs"], row["cv_pct"],
            row["concentration_in_well_ug_mL"], row["final_concentration_ug_mL"], row["qc_flags"],
        ]
        for ci, val in enumerate(vals, 1):
            flag_this = (ci == 10 and flagged) or (ci == 7 and high_cv)
            cell = ws_unk.cell(row=r, column=ci, value=val)
            cell.border    = _THIN
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if ci == 10:
                cell.fill = _FLAG_FILL if flagged else _OK_FILL
            elif flag_this:
                cell.fill = _FLAG_FILL
            # cols 5 (raw abs), 6 (corr abs), 8 (CIW), 9 (final conc) → 9 dp
            if ci in (5, 6, 8, 9) and isinstance(val, float) and not np.isnan(val):
                cell.number_format = _NUM_FMT

    _auto_width(ws_unk)

    # ── Sheet 4: Warnings ────────────────────────────────────────────────────
    ws_w = wb.create_sheet("Warnings")
    ws_w["A1"].value = "QC Warnings"
    ws_w["A1"].font  = Font(size=14, bold=True, color="CC0000")
    ws_w.row_dimensions[1].height = 24

    _header_row(ws_w, 2, ["Warning"])

    warn_list: list[str] = []

    if not np.isnan(r_squared) and r_squared < 0.99:
        warn_list.append(
            f"Low R² = {r_squared:.5f}  (BCA assays should have R² ≥ 0.99 — check standards)"
        )

    for _, row in standards_df.iterrows():
        if float(row["cv_pct"]) > cv_threshold:
            warn_list.append(
                f"Standard {row['concentration_ug_mL']} µg/mL: "
                f"High CV = {row['cv_pct']:.1f}% (threshold: {cv_threshold}%)"
            )

    for _, row in unknowns_results.iterrows():
        if row["qc_flags"] != "OK":
            warn_list.append(f"Sample '{row['sample_name']}': {row['qc_flags']}")

    if not warn_list:
        warn_list.append("No QC warnings — all samples passed.")

    for i, msg in enumerate(warn_list, start=3):
        cell = ws_w.cell(row=i, column=1, value=msg)
        cell.border    = _THIN
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.fill = _OK_FILL if msg.startswith("No QC") else _WARN_FILL
        ws_w.row_dimensions[i].height = 20

    ws_w.column_dimensions["A"].width = 80

    # ── Sheet 5: Western Loading (optional) ─────────────────────────────────
    if western_df is not None and western_params is not None:
        ws_wb = wb.create_sheet("Western_Loading")

        # Parameter summary block
        ws_wb["A1"].value = "Western Blot Loading Parameters"
        ws_wb["A1"].font  = Font(size=13, bold=True, color="2E4057")

        lys_p = western_params["lysate_parts"]
        buf_p = western_params["buffer_parts"]
        cf    = (lys_p + buf_p) / lys_p
        param_rows = [
            ("Sample buffer ratio (lysate:buffer)",  f"{lys_p}:{buf_p}"),
            ("Concentration correction factor",      f"÷{cf:.4f}"),
            ("Target protein per lane (µg)",         western_params["target_ug"]),
            ("Max well volume (µL)",                 western_params.get("max_well_vol") or "—"),
        ]
        for i, (label, value) in enumerate(param_rows, start=2):
            ws_wb.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws_wb.cell(row=i, column=2, value=value)

        # Data table (starts 2 rows after the param block)
        data_start = len(param_rows) + 3

        wb_headers = [
            "Sample Name", "Neat Conc (µg/mL)", "Conc in Sample Buffer (µg/mL)",
            "Target (µg)", "Lysate Vol (µL)", "Sample Buffer Vol (µL)",
            "Total Load Vol (µL)", "WB Flags",
        ]
        _header_row(ws_wb, data_start, wb_headers)

        wb_col_map = [
            "sample_name", "neat_conc_ug_mL", "conc_in_sample_buffer_ug_mL",
            "target_ug", "lysate_vol_uL", "sample_buffer_vol_uL",
            "total_load_vol_uL", "wb_flags",
        ]

        for ri, row in western_df.iterrows():
            r = data_start + ri + 1
            flagged = row["wb_flags"] != "OK"

            for ci, col_key in enumerate(wb_col_map, 1):
                val  = row[col_key]
                cell = ws_wb.cell(row=r, column=ci, value=None if (isinstance(val, float) and np.isnan(val)) else val)
                cell.border    = _THIN
                cell.alignment = Alignment(horizontal="center", vertical="center")

                if ci == 8:  # WB Flags column
                    cell.fill = _FLAG_FILL if flagged else _OK_FILL
                # cols 2 (neat conc), 3 (buf conc), 5 (lysate vol), 6 (buf vol), 7 (total vol) → 9 dp
                if ci in (2, 3, 5, 6, 7) and isinstance(val, float) and not np.isnan(val):
                    cell.number_format = _NUM_FMT

        _auto_width(ws_wb)

    wb.save(output_path)
    print(f"  [+] Excel workbook saved   → {output_path}")


# ═══════════════════════════════════════════════════════════════════════════════
# TERMINAL RESULTS SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════

def print_summary(
    blank_avg: float,
    blank_cv: float,
    standards_df: pd.DataFrame,
    unknowns_results: pd.DataFrame,
    fit_mode: str,
    slope: float,
    intercept: float,
    r_squared: float,
    cv_threshold: float,
) -> None:
    """Print a concise results table to the terminal."""
    sep = "═" * 72

    print("\n" + sep)
    print("  RESULTS SUMMARY")
    print(sep)

    print(f"\n  0 µg/mL std avg abs  : {blank_avg:.5f}   (CV: {blank_cv:.2f}%)")

    print(f"\n  Standard Curve  [{fit_mode}]")
    print(f"  {'─' * 45}")
    if fit_mode == FIT_LINEAR:
        sign = "+" if intercept >= 0 else "−"
        print(f"  Equation  : Conc = {slope:.4f} × Adj.Abs {sign} {abs(intercept):.4f}")
        print(f"  R²        : {r_squared:.6f}", end="")
        if r_squared < 0.99:
            print("  ← WARNING: R² < 0.99", end="")
        print()
    else:
        n_pts = len(standards_df)
        abs_lo = standards_df["blank_corrected_abs"].min()
        abs_hi = standards_df["blank_corrected_abs"].max()
        print(f"  {n_pts} standards used for interpolation")
        print(f"  Absorbance range : {abs_lo:.5f} – {abs_hi:.5f}")

    print(f"\n  Standards (blank-corrected)")
    print(f"  {'─' * 50}")
    print(f"  {'Conc (µg/mL)':>14}  {'Corr. Abs':>11}  {'CV%':>6}  Flag")
    for _, row in standards_df.iterrows():
        flag = "[HIGH CV]" if row["cv_pct"] > cv_threshold else ""
        print(
            f"  {row['concentration_ug_mL']:>14.1f}  "
            f"{row['blank_corrected_abs']:>11.5f}  "
            f"{row['cv_pct']:>5.1f}%  {flag}"
        )

    print(f"\n  Unknown Samples")
    print(f"  {'─' * 72}")
    print(f"  {'Sample':<20}  {'Corr. Abs':>10}  {'Final Conc':>14}  {'CV%':>5}  Flags")
    for _, row in unknowns_results.iterrows():
        conc = (
            f"{row['final_concentration_ug_mL']:.3f} µg/mL"
            if pd.notna(row["final_concentration_ug_mL"])
            else "N/A"
        )
        print(
            f"  {row['sample_name']:<20}  "
            f"{row['blank_corrected_abs']:>10.5f}  "
            f"{conc:>14}  "
            f"{row['cv_pct']:>4.1f}%  "
            f"{row['qc_flags']}"
        )

    print("\n" + sep + "\n")


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def sanitize_name(name: str) -> str:
    """Replace unsafe path characters with underscores and strip extremes."""
    safe = re.sub(r'[^\w\-.]', '_', name.strip())
    safe = re.sub(r'_+', '_', safe).strip('_')
    return safe[:120] if safe else "BCA_Run"


def create_run_folder() -> str:
    """
    Create BCA_Runs/<run_name>/ in the current working directory.
    If the folder already exists, append _YYYYMMDD_HHMMSS before creating.
    Returns the path to the new folder.
    """
    today = datetime.now().strftime("%Y-%m-%d")
    print(f"\n  Enter a name for this BCA run (e.g. {today}_A549_RAB7A_BCA).")
    print("  Press Enter to use the date only.")
    raw = input("  Run name: ").strip()
    run_name = sanitize_name(raw) if raw else f"{today}_BCA_Run"

    bca_runs = os.path.join(os.getcwd(), "BCA_Runs")
    os.makedirs(bca_runs, exist_ok=True)

    run_dir = os.path.join(bca_runs, run_name)
    if os.path.exists(run_dir):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        run_dir = os.path.join(bca_runs, f"{run_name}_{ts}")
        print(f"  [!] Folder already exists — using timestamped name.")

    os.makedirs(run_dir)
    return run_dir


def main() -> None:
    parser = argparse.ArgumentParser(
        prog="bca_calculator.py",
        description="BCA Protein Assay Calculator — analyze plate reader data and export results.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python bca_calculator.py plate_data.xlsx
  python bca_calculator.py plate_data.csv --output results/my_run.xlsx --cv-threshold 10
        """,
    )
    parser.add_argument(
        "input_file",
        help="Path to input CSV or Excel file (.csv / .xls / .xlsx)",
    )
    parser.add_argument(
        "--output", "-o",
        default=None,
        help="Output Excel file (default: <input_name>_results.xlsx)",
    )
    parser.add_argument(
        "--sheet", "-s",
        default=None,
        help="Excel sheet name or index to read (default: first sheet)",
    )
    parser.add_argument(
        "--cv-threshold",
        type=float,
        default=DEFAULT_CV_THRESHOLD,
        help=f"Percent CV above which replicates are flagged (default: {DEFAULT_CV_THRESHOLD})",
    )
    parser.add_argument(
        "--fit-mode",
        choices=[FIT_LINEAR, FIT_PTP],
        default=FIT_LINEAR,
        help=(
            f"Standard curve fit method. '{FIT_LINEAR}' (default) = OLS linear regression. "
            f"'{FIT_PTP}' = piecewise interpolation between adjacent standards "
            "(Pierce protocol recommendation for microplate data)."
        ),
    )

    args = parser.parse_args()

    input_path   = args.input_file
    cv_threshold = args.cv_threshold
    sheet        = args.sheet
    fit_mode     = args.fit_mode

    if not os.path.exists(input_path):
        print(f"\n[ERROR] File not found: {input_path}\n")
        sys.exit(1)

    input_filename = os.path.basename(input_path)
    input_stem     = os.path.splitext(input_filename)[0]

    # ── Header banner ─────────────────────────────────────────────────────────
    print("\n" + "═" * 72)
    print("  BCA PROTEIN ASSAY CALCULATOR")
    print("═" * 72)
    print(f"  Input  : {input_path}")
    print(f"  CV flag threshold: {cv_threshold}%")
    print(f"  Fit mode         : {fit_mode}")
    print("═" * 72)

    # ── Create run folder and route all outputs there ─────────────────────────
    run_dir     = create_run_folder()
    out_name    = os.path.basename(args.output) if args.output else f"{input_stem}_results.xlsx"
    output_excel = os.path.join(run_dir, out_name)
    output_plot  = os.path.join(run_dir, os.path.splitext(out_name)[0] + "_standard_curve.png")

    # Copy the input file into the run folder
    shutil.copy2(input_path, os.path.join(run_dir, input_filename))
    print(f"\n  Run folder : {run_dir}")

    # ── Step 1: Load ──────────────────────────────────────────────────────────
    print("\n[1/7] Loading plate data …")
    try:
        well_data = load_plate_data(input_path, sheet=sheet)
    except Exception as exc:
        print(f"\n[ERROR] Could not load file: {exc}\n")
        sys.exit(1)

    print(f"  Loaded {len(well_data)} wells.")
    display_plate(well_data)

    # ── Step 2: Interactive well assignment ───────────────────────────────────
    print("[2/7] Assign wells\n")
    blank_wells    = ["A9", "B9", "C9"]   # always the 0 µg/mL standard wells
    print(f"  0 µg/mL standard wells : A9, B9, C9 (fixed)")
    standard_defs  = [
        {"wells": ["A1","B1","C1"], "concentration": 2000},
        {"wells": ["A2","B2","C2"], "concentration": 1500},
        {"wells": ["A3","B3","C3"], "concentration": 1000},
        {"wells": ["A4","B4","C4"], "concentration":  750},
        {"wells": ["A5","B5","C5"], "concentration":  500},
        {"wells": ["A6","B6","C6"], "concentration":  250},
        {"wells": ["A7","B7","C7"], "concentration":  125},
        {"wells": ["A8","B8","C8"], "concentration":   25},
    ]
    print("  Standards (A1:C8)      : 2000, 1500, 1000, 750, 500, 250, 125, 25 µg/mL (fixed)\n")
    lysate_parts, buffer_parts = 5, 1     # always 5:1
    dilution_factor = (lysate_parts + buffer_parts) / lysate_parts  # 1.2
    unknown_defs   = prompt_unknowns(well_data)
    for u in unknown_defs:
        u["dilution"] = dilution_factor

    # ── Step 3: Blank correction & replicate statistics ───────────────────────
    print("\n[3/7] Computing blank correction and replicate statistics …")
    blank_avg, blank_cv, standards_df, unknowns_df = compute_blank_corrected(
        well_data, blank_wells, standard_defs, unknown_defs
    )
    print(f"  Blank mean: {blank_avg:.5f}  CV: {blank_cv:.2f}%")

    # ── Step 4: Fit standard curve ────────────────────────────────────────────
    slope = intercept = r_squared = np.nan
    if fit_mode == FIT_LINEAR:
        print("\n[4/7] Fitting linear standard curve …")
        slope, intercept, r_squared = fit_linear_curve(standards_df)
        print(f"  Slope={slope:.7f}  Intercept={intercept:.7f}  R²={r_squared:.6f}")
        if r_squared < 0.99:
            print("  [!] WARNING: R² < 0.99 — the standard curve fit is poor.")
    else:
        print("\n[4/7] Using point-to-point interpolation (Pierce protocol recommendation) …")
        std_sorted = standards_df.sort_values("blank_corrected_abs")
        abs_lo = std_sorted["blank_corrected_abs"].iloc[0]
        abs_hi = std_sorted["blank_corrected_abs"].iloc[-1]
        print(f"  {len(standards_df)} standards, absorbance range {abs_lo:.5f} – {abs_hi:.5f}")

    # ── Step 5: Calculate unknown concentrations ──────────────────────────────
    print("\n[5/7] Calculating unknown protein concentrations …")
    std_min = standards_df["concentration_ug_mL"].min()
    std_max = standards_df["concentration_ug_mL"].max()
    unknowns_results = calculate_concentrations(
        unknowns_df, fit_mode, standards_df, slope, intercept, std_min, std_max, cv_threshold
    )

    # Print terminal summary
    print_summary(
        blank_avg, blank_cv, standards_df, unknowns_results,
        fit_mode, slope, intercept, r_squared, cv_threshold,
    )

    # ── Step 5b: Western blot loading calculator (always runs at 10 µg/lane) ──
    western_params = {
        "target_ug":    10.0,
        "max_well_vol": None,
        "lysate_parts": lysate_parts,
        "buffer_parts": buffer_parts,
    }
    western_df = calculate_western_loading(unknowns_results, western_params)
    print_western_table(western_df, western_params)

    # ── Step 6: Generate plot ─────────────────────────────────────────────────
    print("[6/7] Generating standard curve plot …")
    generate_standard_curve_plot(standards_df, fit_mode, slope, intercept, r_squared, output_plot)

    # ── Step 7: Export Excel ──────────────────────────────────────────────────
    print("\n[7/7] Exporting results to Excel …")
    export_to_excel(
        output_excel, blank_avg, blank_cv, blank_wells,
        standards_df, unknowns_results,
        fit_mode, slope, intercept, r_squared, cv_threshold, input_filename,
        western_df=western_df, western_params=western_params,
    )

    print("\n" + "═" * 72)
    print("  Done.  All files saved to run folder:")
    print(f"  {run_dir}")
    print()
    print(f"  Input  (copy) : {os.path.join(run_dir, input_filename)}")
    print(f"  Excel         : {output_excel}")
    print(f"  Plot          : {output_plot}")
    print("═" * 72 + "\n")


if __name__ == "__main__":
    main()
