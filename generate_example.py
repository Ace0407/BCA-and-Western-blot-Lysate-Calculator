#!/usr/bin/env python3
"""
generate_example.py — Creates a realistic example BCA plate reader input file.

The plate layout produced:

  Row A, cols 1-3  : Blanks (3 replicates)
  Row B, cols 1-2  : BSA standard   25 µg/mL
  Row C, cols 1-2  : BSA standard  125 µg/mL
  Row D, cols 1-2  : BSA standard  250 µg/mL
  Row E, cols 1-2  : BSA standard  500 µg/mL
  Row F, cols 1-2  : BSA standard  750 µg/mL
  Row G, cols 1-2  : BSA standard 1000 µg/mL
  Row H, cols 1-2  : BSA standard 2000 µg/mL
  Row A, cols 5-6  : Unknown — Lysate_A  (dilution ×2)
  Row B, cols 5-6  : Unknown — Lysate_B  (dilution ×5)
  Row C, cols 5-6  : Unknown — Membrane_Prep  (dilution ×1)
  Row D, cols 5-6  : Unknown — Serum_Fraction  (dilution ×10) ← intentionally above curve
  Row E, cols 5-6  : Unknown — Dilute_Sample  (dilution ×1)  ← intentionally below curve
  Row F, cols 5-6  : Unknown — HighCV_Sample  (high replicate noise)

Usage:
    python generate_example.py
    → writes example_plate.xlsx in the current directory
    → prints a hint showing how to enter wells when you run bca_calculator.py
"""

import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

random.seed(42)

# ── True underlying parameters (the "real" standard curve) ───────────────────
TRUE_SLOPE     = 0.001540   # absorbance per µg/mL
TRUE_INTERCEPT = 0.0        # blank-corrected, so true intercept near zero
BLANK_ABS      = 0.0985     # mean blank absorbance
NOISE_STD      = 0.0025     # typical pipetting noise (small, ±0.25%)

PLATE_ROWS = list("ABCDEFGH")
N_COLS     = 12


def sim(conc: float, noise: float = NOISE_STD) -> float:
    """Simulate an absorbance reading at a given protein concentration."""
    true_abs = BLANK_ABS + TRUE_SLOPE * conc
    return round(true_abs + random.gauss(0, noise), 4)


def build_well_data() -> dict[str, float]:
    """Return a dict mapping well-ID → absorbance."""
    data: dict[str, float] = {}

    # Blanks: A1, A2, A3
    for col in [1, 2, 3]:
        data[f"A{col}"] = round(BLANK_ABS + random.gauss(0, NOISE_STD), 4)

    # BSA standards (duplicate wells)
    standards = [
        ("B", 25),
        ("C", 125),
        ("D", 250),
        ("E", 500),
        ("F", 750),
        ("G", 1000),
        ("H", 2000),
    ]
    for row, conc in standards:
        data[f"{row}1"] = sim(conc)
        data[f"{row}2"] = sim(conc)

    # Unknown samples (columns 5-6)
    unknowns = [
        ("A", 320.0),    # Lysate_A  — in-range, dilution ×2 → real=640 µg/mL
        ("B", 185.0),    # Lysate_B  — in-range, dilution ×5 → real=925 µg/mL
        ("C", 410.0),    # Membrane_Prep — in-range, dilution ×1
        ("D", 2800.0),   # Serum_Fraction — ABOVE CURVE (>2000 µg/mL)
        ("E", 10.0),     # Dilute_Sample — BELOW CURVE (<25 µg/mL)
        ("F", 500.0),    # HighCV_Sample — high noise
    ]
    high_cv_noise = 0.020  # deliberately high noise for sample F

    for row, conc in unknowns:
        noise = high_cv_noise if row == "F" else NOISE_STD
        data[f"{row}5"] = sim(conc, noise)
        data[f"{row}6"] = sim(conc, noise)

    return data


def write_excel(well_data: dict[str, float], output_path: str) -> None:
    """Write the plate grid to an Excel file (plate-reader grid format)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plate"

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    hdr_fill   = PatternFill("solid", fgColor="2E4057")
    hdr_font   = Font(color="FFFFFF", bold=True, size=11)
    label_font = Font(bold=True)

    # Column headers: blank + 1-12
    ws.cell(row=1, column=1, value="").fill   = hdr_fill
    ws.cell(row=1, column=1).border           = thin
    for col in range(1, N_COLS + 1):
        cell = ws.cell(row=1, column=col + 1, value=col)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal="center")
        cell.border    = thin
    ws.row_dimensions[1].height = 20
    ws.column_dimensions["A"].width = 5

    # Row headers A-H + data
    for ri, row_label in enumerate(PLATE_ROWS, start=2):
        lbl = ws.cell(row=ri, column=1, value=row_label)
        lbl.fill      = hdr_fill
        lbl.font      = hdr_font
        lbl.alignment = Alignment(horizontal="center")
        lbl.border    = thin

        for col in range(1, N_COLS + 1):
            well_id = f"{row_label}{col}"
            val     = well_data.get(well_id)
            cell    = ws.cell(row=ri, column=col + 1, value=val)
            cell.border    = thin
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = 9

    wb.save(output_path)
    print(f"[+] Example input file written → {output_path}")


def print_usage_hint(well_data: dict[str, float]) -> None:
    """Print a ready-to-paste guide for bca_calculator.py."""
    sep = "─" * 70
    print()
    print(sep)
    print("  HOW TO USE THIS FILE WITH bca_calculator.py")
    print(sep)
    print("  Run:")
    print("    python bca_calculator.py example_plate.xlsx")
    print()
    print("  When prompted, enter the following:")
    print()
    print("  Blanks:     A1 A2 A3")
    print()
    print("  Standards:")
    print("    #1  wells: B1 B2   conc: 25")
    print("    #2  wells: C1 C2   conc: 125")
    print("    #3  wells: D1 D2   conc: 250")
    print("    #4  wells: E1 E2   conc: 500")
    print("    #5  wells: F1 F2   conc: 750")
    print("    #6  wells: G1 G2   conc: 1000")
    print("    #7  wells: H1 H2   conc: 2000")
    print("    (press Enter to stop)")
    print()
    print("  Unknowns (sample name / wells / dilution factor):")
    print("    #1  wells: A5 A6   name: Lysate_A       dilution: 2")
    print("    #2  wells: B5 B6   name: Lysate_B       dilution: 5")
    print("    #3  wells: C5 C6   name: Membrane_Prep  dilution: 1")
    print("    #4  wells: D5 D6   name: Serum_Fraction dilution: 10")
    print("    #5  wells: E5 E6   name: Dilute_Sample  dilution: 1")
    print("    #6  wells: F5 F6   name: HighCV_Sample  dilution: 1")
    print("    (press Enter to stop)")
    print()
    print("  Expected QC flags:")
    print("    Serum_Fraction  → ABOVE CURVE RANGE")
    print("    Dilute_Sample   → BELOW CURVE RANGE")
    print("    HighCV_Sample   → HIGH CV (intentional noise)")
    print(sep)
    print()


if __name__ == "__main__":
    output = "example_plate.xlsx"
    well_data = build_well_data()
    write_excel(well_data, output)
    print_usage_hint(well_data)
